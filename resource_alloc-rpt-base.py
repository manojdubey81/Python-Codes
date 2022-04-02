#**********************************************************************************************************************************************************
#       IMPORT LIBRARIES
#**********************************************************************************************************************************************************
import os
import xlrd
import pandas as pd
from pandasql import sqldf 
from datetime import datetime, date
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
#**********************************************************************************************************************************************************
 


#**********************************************************************************************************************************************************
#   DIALOG WINDOW BOX HEADING AT TOP LEFT CORNER
#**********************************************************************************************************************************************************

root = Tk()
root.title("CAMMIS Resource Projection Script")


#**********************************************************************************************************************************************************
#   LABEL TO DISPLAY "Resource Master File" BEFORE BOX BUTTON
#   BOX BUTTON TO SELECT "Resource Master File"
#   LABEL FOR DISPLAY SELECTED RESOURCE MASTER FILE NAME
#**********************************************************************************************************************************************************


file_label1 = ttk.Label(root, text="Resource Master File:").grid(row=1, column=0, padx=10, pady=10)

open_resource_file = ttk.Button(text="Browse files", command = lambda:read_resource_master())
open_resource_file.grid(row=1, column=1, columnspan=2, padx=10, pady=10)

file_label2 = ttk.Label(root, text="Selected File Is:").grid(row=2, column=0, padx=20, pady=20)
box1_path = ttk.Label(root, text="No File Selected Yet")
box1_path.grid(row=2, column=1, columnspan=4, padx=20, pady=20)

#**********************************************************************************************************************************************************
#   LABEL TO DISPLAY "Practitioners Forcast File" BEFORE BOX BUTTON
#   BOX BUTTON TO SELECT "Practitioners Raw Data File"
#   LABEL TO DISLAY SELECTED PRACTITIONERS FILE NAME
#**********************************************************************************************************************************************************

file_label4 = ttk.Label(root, text="Practitioners Forcast File:").grid(row=3, column=0, padx=10, pady=10)

open_pract_file = ttk.Button(text="Browse Files", command = lambda:read_raw_data())
open_pract_file.grid(row=3, column=1, columnspan=2, padx=10, pady=10)

file_label5 = ttk.Label(root, text="Selected File Is:").grid(row=4, column=0, padx=20, pady=20)
box2_path = ttk.Label(root, text="No File Selected Yet")
box2_path.grid(row=4, column=1, columnspan=4, padx=20, pady=20)

#**********************************************************************************************************************************************************
#   LABEL FOR "Click on Start Button" to START THE PROCESSING
#**********************************************************************************************************************************************************

file_label7 = ttk.Label(root, text="Click Start to Proceed:").grid(row=5, column=0, padx=20, pady=20)

stare_scripte = ttk.Button(root, text="Start", command = lambda:start_app())
stare_scripte.grid(row=5, column=1, columnspan=2, padx=60, pady=60)


#**********************************************************************************************************************************************************
#       PANDAS SQL
#**********************************************************************************************************************************************************

mysql = lambda q: sqldf(q, globals())



read_raw_data = ''
read_resource_master = ''

pd.set_option("display.max_rows", None, "display.max_columns", None)

#**********************************************************************************************************************************************************
#       USER DEFINED FUCTIONS
#**********************************************************************************************************************************************************

#**********************************************************************************************************************************************************
#   ROW COUNT FUNTION IF NO ROW SELECTED IT WILL ADD 1 to num to HIGHLIGHT 1 row
#**********************************************************************************************************************************************************

def retd(num):
    if num == 0:
        num  += 1
    return num

#**********************************************************************************************************************************************************
#   SkillCategory USER DEFINED FUCTION
#**********************************************************************************************************************************************************

def SkillCategory(SkillName):

        SkillNameUpper = SkillName.upper()
        skills = "MAINFRAME, DOT NET, JAVA, DBA, PM, SME, BA, DOC SPL"
        if skills.find(SkillNameUpper) >= 0:
                SkillCategory = SkillNameUpper
        else:
                SkillCategory = "OTHER"

        return SkillCategory

#**********************************************************************************************************************************************************
#   Skillname USER DEFINED FUNCTION
#**********************************************************************************************************************************************************

def skillname(ResourceName):

        ResourceNameUpper = ResourceName.upper()
        if ResourceNameUpper.find('.ARCH') >= 0:
                SkillName = "ARCHITECT"
        elif ResourceNameUpper.find('.BO') >= 0:
                SkillName = "BO"
        elif ResourceNameUpper.find('.DB') >= 0:
                SkillName = "DBA"
        elif ResourceNameUpper.find('.NET') >= 0:
                SkillName = "DOT NET"
        elif ResourceNameUpper.find('.VB') >= 0:
                SkillName = "DOT NET"
        elif ResourceNameUpper.find('.ETL') >= 0:
                SkillName = "ETL"
        elif ResourceNameUpper.find('.JAVA') >= 0:
                SkillName = "JAVA"
        elif ResourceNameUpper.find('.MF') >= 0:
                SkillName = "MAINFRAME"
        elif ResourceNameUpper.find('.PM') >= 0:
                SkillName = "PM"
        elif ResourceNameUpper.find('.SME') >= 0:
                SkillName = "SME"
        elif ResourceNameUpper.find('.SURGE') >= 0:
                SkillName = "SURGE"
        elif ResourceNameUpper.find('.USOFT') >= 0:
                SkillName = "USOFT"
        elif ResourceNameUpper.find('.RAIS') >= 0:
                SkillName = "USOFT"
        elif ResourceNameUpper.find('.ORACLEDBA') >= 0:
                SkillName = "ORACLE DBA"
        elif ResourceNameUpper.find('.MCWEB') >= 0:
                SkillName = "DOT NET"
        elif ResourceNameUpper.find('.CMS64') >= 0:
                SkillName = "DOT NET"
        elif ResourceNameUpper.find('.WTX') >= 0:
                SkillName = "WTX"
        elif ResourceNameUpper.find('.ITX') >= 0:
                SkillName = "WTX"
        elif ResourceNameUpper.find('.BA') >= 0:
                SkillName = "BA"
        elif ResourceNameUpper.find('.DOC SPL') >= 0:
                SkillName = "DOC SPL"
        elif ResourceNameUpper.find('.BUSINESS CONSULTANT') >= 0:
                SkillName = "SME"
        else:
                SkillName = "OTHERS"

        return SkillName

#**********************************************************************************************************************************************************
#   DIALOG BOX CLOSING FUNCTION 
#**********************************************************************************************************************************************************
                    
def close_dialog_box():
        root.destroy()
        

#**********************************************************************************************************************************************************
#   THIS FUNCTION WILL SELECT RESOURCE MASTER AND DISPLAY FILE NAME NEXT TO BUTTON
#**********************************************************************************************************************************************************

def read_resource_master():

        global read_resource_master
        global Resource_Master
        global Resource_Master
        
        r_fl = filedialog.askopenfilename(initialdir="/", filetypes=(("CSV File", ".*csv"), ("All File", "*.*")))
       
        if r_fl:
                try:
                        Resource_Master = pd.read_csv(r_fl, names=["SNo", "ResourceName", "Skill", "Type", "Status", "Team"])
                        r_fl_name = os.path.basename(r_fl)
                        box1_path["text"] = r_fl_name
                        read_resource_master = 'y'
                
                except FileNotFoundError:
                        messagebox.showerror("Information" , "File Not Found!")
                        box1_path["text"] = 'Please Select Correct File'
                        return None
                
                except ValueError:
                        messagebox.showerror("Information" , "The file you have choosen is invalid")
                        box1_path["text"] = 'Please Select Correct File'
                        return None        
                
                
#**********************************************************************************************************************************************************
#   THIS FUNCTION TO SELECT PRACTITIONERS FORECAST FILE AND DISPLAY THE SELECTED FILE NAME NEXT TO BUTTON
#**********************************************************************************************************************************************************

def read_raw_data():
        
        global read_raw_data
        global dt_in_ccyymmdd
        global Raw_Data
        
        p_fl = filedialog.askopenfilename(initialdir="/", filetypes=(("Excel File", ".*xls*"), ("All File", "*.*")))

        if p_fl:
                try:
                        p_fl = r"{}".format(p_fl)
                        Raw_Data = pd.read_excel(p_fl, sheet_name="Raw Data")
                        Raw_Data.rename({'Weekend Date': 'WeekendDate', 'Resource Name': 'ResourceName'}, axis=1, inplace=True)
                        Raw_Data['WeekendDate'] = Raw_Data['WeekendDate'].dt.strftime('%m/%d/%Y')
                        p_fl_name = os.path.basename(p_fl)
                        box2_path["text"] = p_fl_name

                        p_fl_name_split = p_fl_name.split()
                        dt_frm_inp_fl   = p_fl_name_split[2]
                        dt_in_ccyymmdd  = dt_frm_inp_fl[6:] + dt_frm_inp_fl[:2] + dt_frm_inp_fl[3:5]

                        read_raw_data = 'y'
                        

                except FileNotFoundError:
                        messagebox.showerror("Information" , "File Not Found!")
                        box2_path["text"] = 'Please Select Correct File'
                        return None
                
                except ValueError:
                        messagebox.showerror("Information" , "The file you have choosen is invalid")
                        box2_path["text"] = 'Please Select Correct File'
                        return None        
        


#**********************************************************************************************************************************************************
#   WRITE OUTPUT EXCEL FILE
#**********************************************************************************************************************************************************

def create_output():

        style = TableStyleInfo(name="TableStyleMedium4", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)


        with pd.ExcelWriter("Projections Report - gen_" + dt_in_ccyymmdd + "_SDN.xlsx", engine='openpyxl') as writer:

#**********************************************************************************************************************************************************
# OUTPUT EXCEL DATA WRITTEN STARTS HERE INTO DIFFERENT WORKSHEET WITH APPROPRIATE TABLE HEADING
#**********************************************************************************************************************************************************

                wrk_sheet1 = 'Overall Projections'
                wrk_sheet2 = 'Report of Unallocated Resources'
                wrk_sheet3 = 'Approved Actuals by Skill'
                wrk_sheet4 = 'Total Vacations By Skill'
                wrk_sheet5 = 'Total Submitted By Skill'
                wrk_sheet6 = 'Total Placeholders by skill'
                wrk_sheet7 = 'Covered Resource Counts'

                table_1_aa.to_excel(writer, sheet_name=wrk_sheet3, startrow = 1, startcol=0, index=False)
                table_2_aa.to_excel(writer, sheet_name=wrk_sheet3, startrow = 1, startcol=3, index=False)
                table_3_aa.to_excel(writer, sheet_name=wrk_sheet3, startrow = 1, startcol=7, index=False)
                table_4_aa.to_excel(writer, sheet_name=wrk_sheet3, startrow = 1, startcol=11, index=False)
                table_5_aa.to_excel(writer, sheet_name=wrk_sheet3, startrow = 1, startcol=15, index=False)
                table_6_aa.to_excel(writer, sheet_name=wrk_sheet3, startrow = 1, startcol=19, index=False)
                table_7_aa.to_excel(writer, sheet_name=wrk_sheet3, startrow = 1, startcol=23, index=False)
                table_8_aa.to_excel(writer, sheet_name=wrk_sheet3, startrow = 1, startcol=27, index=False)
                table_9_aa.to_excel(writer, sheet_name=wrk_sheet3, startrow = 1, startcol=31, index=False)
                table_10_aa.to_excel(writer, sheet_name=wrk_sheet3, startrow = 1, startcol=35, index=False)

                table_1_tv.to_excel(writer, sheet_name=wrk_sheet4, startrow = 1, startcol=0, index=False)
                table_2_tv.to_excel(writer, sheet_name=wrk_sheet4, startrow = 1, startcol=3, index=False)
                table_3_tv.to_excel(writer, sheet_name=wrk_sheet4, startrow = 1, startcol=7, index=False)
                table_4_tv.to_excel(writer, sheet_name=wrk_sheet4, startrow = 1, startcol=11, index=False)
                table_5_tv.to_excel(writer, sheet_name=wrk_sheet4, startrow = 1, startcol=15, index=False)
                table_6_tv.to_excel(writer, sheet_name=wrk_sheet4, startrow = 1, startcol=19, index=False)
                table_7_tv.to_excel(writer, sheet_name=wrk_sheet4, startrow = 1, startcol=23, index=False)
                table_8_tv.to_excel(writer, sheet_name=wrk_sheet4, startrow = 1, startcol=27, index=False)
                table_9_tv.to_excel(writer, sheet_name=wrk_sheet4, startrow = 1, startcol=31, index=False)
                table_10_tv.to_excel(writer, sheet_name=wrk_sheet4, startrow = 1, startcol=35, index=False)

                table_1_ts.to_excel(writer, sheet_name=wrk_sheet5, startrow = 1, startcol=0, index=False)
                table_2_ts.to_excel(writer, sheet_name=wrk_sheet5, startrow = 1, startcol=3, index=False)
                table_3_ts.to_excel(writer, sheet_name=wrk_sheet5, startrow = 1, startcol=7, index=False)
                table_4_ts.to_excel(writer, sheet_name=wrk_sheet5, startrow = 1, startcol=11, index=False)
                table_5_ts.to_excel(writer, sheet_name=wrk_sheet5, startrow = 1, startcol=15, index=False)
                table_6_ts.to_excel(writer, sheet_name=wrk_sheet5, startrow = 1, startcol=19, index=False)
                table_7_ts.to_excel(writer, sheet_name=wrk_sheet5, startrow = 1, startcol=23, index=False)
                table_8_ts.to_excel(writer, sheet_name=wrk_sheet5, startrow = 1, startcol=27, index=False)
                table_9_ts.to_excel(writer, sheet_name=wrk_sheet5, startrow = 1, startcol=31, index=False)
                table_10_ts.to_excel(writer, sheet_name=wrk_sheet5, startrow = 1, startcol=35, index=False)                
                                
                table_1_tp.to_excel(writer, sheet_name=wrk_sheet6, startrow = 1, startcol=0, index=False)
                table_2_tp.to_excel(writer, sheet_name=wrk_sheet6, startrow = 1, startcol=3, index=False)
                table_3_tp.to_excel(writer, sheet_name=wrk_sheet6, startrow = 1, startcol=7, index=False)
                table_4_tp.to_excel(writer, sheet_name=wrk_sheet6, startrow = 1, startcol=11, index=False)
                table_5_tp.to_excel(writer, sheet_name=wrk_sheet6, startrow = 1, startcol=15, index=False)
                table_6_tp.to_excel(writer, sheet_name=wrk_sheet6, startrow = 1, startcol=19, index=False)
                table_7_tp.to_excel(writer, sheet_name=wrk_sheet6, startrow = 1, startcol=23, index=False)
                table_8_tp.to_excel(writer, sheet_name=wrk_sheet6, startrow = 1, startcol=27, index=False)
                table_9_tp.to_excel(writer, sheet_name=wrk_sheet6, startrow = 1, startcol=31, index=False)
                table_10_tp.to_excel(writer, sheet_name=wrk_sheet6, startrow = 1, startcol=35, index=False)

                table_1_cr.to_excel(writer, sheet_name=wrk_sheet7, startrow = 1, startcol=0, index=False)
                table_2_cr.to_excel(writer, sheet_name=wrk_sheet7, startrow = 1, startcol=3, index=False)
                table_3_cr.to_excel(writer, sheet_name=wrk_sheet7, startrow = 1, startcol=6, index=False)
                table_4_cr.to_excel(writer, sheet_name=wrk_sheet7, startrow = 1, startcol=9, index=False)
                table_5_cr.to_excel(writer, sheet_name=wrk_sheet7, startrow = 1, startcol=12, index=False)
                table_6_cr.to_excel(writer, sheet_name=wrk_sheet7, startrow = 1, startcol=15, index=False)
                table_7_cr.to_excel(writer, sheet_name=wrk_sheet7, startrow = 1, startcol=18, index=False)
                table_8_cr.to_excel(writer, sheet_name=wrk_sheet7, startrow = 1, startcol=21, index=False)
                table_9_cr.to_excel(writer, sheet_name=wrk_sheet7, startrow = 1, startcol=24, index=False)
                table_10_cr.to_excel(writer, sheet_name=wrk_sheet7, startrow = 1, startcol=27, index=False)
                table_11_cr.to_excel(writer, sheet_name=wrk_sheet7, startrow = 1, startcol=30, index=False)
                
                
                wb  = writer.book
                wb.create_sheet(index=0, title=wrk_sheet1)
                wb.create_sheet(index=1, title=wrk_sheet2)
               # worksheet1 = writer.sheets[wrk_sheet1]
               # worksheet2 = writer.sheets[wrk_sheet2]
                worksheet3 = writer.sheets[wrk_sheet3]
                worksheet4 = writer.sheets[wrk_sheet4]
                worksheet5 = writer.sheets[wrk_sheet5]
                worksheet6 = writer.sheets[wrk_sheet6]
                worksheet7 = writer.sheets[wrk_sheet7]
                
#**********************************************************************************************************************************************************
#   TABLE HEADERS FOR WORKSHEET3
#**********************************************************************************************************************************************************

                table_1_aa_header = 'Table__1_Approved_Hours_For_Actuals'
                table_2_aa_header = 'Table__2_Approved_Mainframe_Hours_For_Actuals'
                table_3_aa_header = 'Table__3_Approved_DotNet_Hours_For_Actuals'
                table_4_aa_header = 'Table__4_Approved_Java_Hours_For_Actuals'
                table_5_aa_header = 'Table__5_Approved_Dba_Hours_For_Actuals'
                table_6_aa_header = 'Table__6_Approved_Pm_Hours_For_Actuals'
                table_7_aa_header = 'Table__7_Approved_Pgm_Hours_For_Actuals'
                table_8_aa_header = 'Table__8_Approved_Other_Hours_For_Actuals'
                table_9_aa_header = 'Table__9_Approved_Ba_Hours_For_Actuals'
                table_10_aa_header = 'Table__10_Approved_Doc_Hours_For_Actuals'

#**********************************************************************************************************************************************************
#   TABLE HEADERS FOR WORKSHEET4
#**********************************************************************************************************************************************************

                table_1_tv_header = 'Table__1_Vacation_Hours_For_Actuals'
                table_2_tv_header = 'Table__2_Vacation_Mainframe_Hours_For_Actuals'
                table_3_tv_header = 'Table__3_Vacation_DotNet_Hours_For_Actuals'
                table_4_tv_header = 'Table__4_Vacation_Java_Hours_For_Actuals'
                table_5_tv_header = 'Table__5_Vacation_Dba_Hours_For_Actuals'
                table_6_tv_header = 'Table__6_Vacation_Pm_Hours_For_Actuals'
                table_7_tv_header = 'Table__7_Vacation_Sme_Hours_For_Actuals'
                table_8_tv_header = 'Table__8_Vacation_Other_Hours_For_Actuals'
                table_9_tv_header = 'Table__9_Vacation_Ba_Hours_For_Actuals'
                table_10_tv_header = 'Table__10_Vacation_Doc_Hours_For_Actuals'

#**********************************************************************************************************************************************************
#   TABLE HEADERS FOR WORKSHEET5
#**********************************************************************************************************************************************************

                table_1_ts_header = 'Table__1_Submitted_Hours_For_Actuals'
                table_2_ts_header = 'Table__2_Submitted_Mainframe_Hours_For_Actuals'
                table_3_ts_header = 'Table__3_Submitted_DotNet_Hours_For_Actuals'
                table_4_ts_header = 'Table__4_Submitted_Java_Hours_For_Actuals'
                table_5_ts_header = 'Table__5_Submitted_Dba_Hours_For_Actuals'
                table_6_ts_header = 'Table__6_Submitted_Pm_Hours_For_Actuals'
                table_7_ts_header = 'Table__7_Submitted_Sme_Hours_For_Actuals'
                table_8_ts_header = 'Table__8_Submitted_Other_Hours_For_Actuals'
                table_9_ts_header = 'Table__9_Submitted_Ba_Hours_For_Actuals'
                table_10_ts_header = 'Table__10_Submitted_Doc_Hours_For_Actuals'

#**********************************************************************************************************************************************************
#   TABLE HEADERS FOR WORKSHEET6
#**********************************************************************************************************************************************************
                
                table_1_tp_header = 'Table__1_Total_Hours_For_Placeholders'
                table_2_tp_header = 'Table__2_Total_Mainframe_Hours_For_Placeholders'
                table_3_tp_header = 'Table__3_Total_DotNet_Hours_For_Placeholders'
                table_4_tp_header = 'Table__4_Total_Java_Hours_For_Placeholders'
                table_5_tp_header = 'Table__5_Total_Dba_Hours_For_Placeholders'
                table_6_tp_header = 'Table__6_Total_Pm_Hours_For_Placeholders'
                table_7_tp_header = 'Table__7_Total_Pgm_Hours_For_Placeholders'
                table_8_tp_header = 'Table__8_Total_Other_Hours_For_Placeholders'
                table_9_tp_header = 'Table__9_Total_Ba_Hours_For_Placeholders'
                table_10_tp_header = 'Table__10_Total_Doc_Hours_For_Placeholders'

#**********************************************************************************************************************************************************
#   TABLE HEADERS FOR WORKSHEET7
#**********************************************************************************************************************************************************

                table_1_cr_header = 'Table__1_Covered_Resource_Count'
                table_2_cr_header = 'Table__2_Covered_Mainframe_Resource_Count'
                table_3_cr_header = 'Table__3_Covered_DotNet_Resource_Count'
                table_4_cr_header = 'Table__4_Covered_Java_Resource_Count'
                table_5_cr_header = 'Table__5_Covered_Dba_Resource_Count'
                table_6_cr_header = 'Table__6_Covered_Pm_Resource_Count'
                table_7_cr_header = 'Table__7_Covered_Sme_Resource_Count'
                table_8_cr_header = 'Table__8_Covered_Other_Resource_Count'
                table_9_cr_header = 'Table__9_Covered_Ba_Resource_Count'
                table_10_cr_header = 'Table__10_Covered_Doc_Resource_Count'
                table_11_cr_header = 'Table__11_Resource_Master_Counts_By_Type'


# COLOR CODES:
#   '#FFEB9C' -> Dark Yellow, '#FFC7CE' -> Dark Red, '#C6EFCE' -> Dark Green,  '#000000' -> Black,  '#0000FF' -> Blue,  '#800000' -> Brown
#   '#00FFFF' -> Cyan, '#808080' -> Gray,  '#008000' -> Green,  '#00FF00' -> Lime,  '#FF00FF' -> Magenta,  '#000080' -> Navy,  '#FF6600' -> Orange
#   '#FF00FF' -> Pink,  '#800080' -> Purple,  '#FF0000' -> Red,  '#C0C0C0' -> Silver,  '#FFFFFF' -> White,  '#FFFF00' -> Yellow


#**********************************************************************************************************************************************************
#   TABLE's FORMITTING AND AUTO CELL ALIGNMENT FOR WORKSHEET3
#**********************************************************************************************************************************************************

                for idx, col in enumerate(worksheet3, 1):
                        worksheet3.column_dimensions[get_column_letter(idx)].auto_size = True


                (max_row, max_col) = table_1_aa.shape
                max_row = retd(max_row)
                
                cell_range = "A2:B" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_1_aa_header)
                tab.tableStyleInfo = style
                worksheet3.add_table(tab)

                
                (max_row, max_col) = table_2_aa.shape
                max_row = retd(max_row)
                cell_range = "D2:F" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_2_aa_header)
                tab.tableStyleInfo = style
                worksheet3.add_table(tab)

                (max_row, max_col) = table_3_aa.shape
                max_row = retd(max_row)
                cell_range = "H2:J" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_3_aa_header)
                tab.tableStyleInfo = style
                worksheet3.add_table(tab)        

                (max_row, max_col) = table_4_aa.shape
                max_row = retd(max_row)
                cell_range = "L2:N" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_4_aa_header)
                tab.tableStyleInfo = style
                worksheet3.add_table(tab)        

                (max_row, max_col) = table_5_aa.shape
                max_row = retd(max_row)
                cell_range = "P2:R" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_5_aa_header)
                tab.tableStyleInfo = style
                worksheet3.add_table(tab)

                (max_row, max_col) = table_6_aa.shape
                max_row = retd(max_row)
                cell_range = "T2:V" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_6_aa_header)
                tab.tableStyleInfo = style
                worksheet3.add_table(tab)
                
                (max_row, max_col) = table_7_aa.shape
                max_row = retd(max_row)
                cell_range = "X2:Z" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_7_aa_header)
                tab.tableStyleInfo = style
                worksheet3.add_table(tab)

                (max_row, max_col) = table_8_aa.shape
                max_row = retd(max_row)
                cell_range = "AB2:AD" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_8_aa_header)
                tab.tableStyleInfo = style
                worksheet3.add_table(tab)

                (max_row, max_col) = table_9_aa.shape
                max_row = retd(max_row)
                cell_range = "AF2:AH" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_9_aa_header)
                tab.tableStyleInfo = style
                worksheet3.add_table(tab)

                (max_row, max_col) = table_10_aa.shape
                max_row = retd(max_row)
                cell_range = "AJ2:AL" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_10_aa_header)
                tab.tableStyleInfo = style
                worksheet3.add_table(tab)

#**********************************************************************************************************************************************************
#   TABLE's FORMITTING AND AUTO CELL ALIGNMENT FOR WORKSHEET4
#**********************************************************************************************************************************************************


                for idx, col in enumerate(worksheet4, 1):
                        worksheet4.column_dimensions[get_column_letter(idx)].auto_size = True


                (max_row, max_col) = table_1_tv.shape
                max_row = retd(max_row)
                
                cell_range = "A2:B" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_1_tv_header)
                tab.tableStyleInfo = style
                worksheet4.add_table(tab)

                
                (max_row, max_col) = table_2_tv.shape
                max_row = retd(max_row)
                cell_range = "D2:F" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_2_tv_header)
                tab.tableStyleInfo = style
                worksheet4.add_table(tab)

                (max_row, max_col) = table_3_tv.shape
                max_row = retd(max_row)
                cell_range = "H2:J" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_3_tv_header)
                tab.tableStyleInfo = style
                worksheet4.add_table(tab)        

                (max_row, max_col) = table_4_tv.shape
                max_row = retd(max_row)
                cell_range = "L2:N" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_4_tv_header)
                tab.tableStyleInfo = style
                worksheet4.add_table(tab)        

                (max_row, max_col) = table_5_tv.shape
                max_row = retd(max_row)
                cell_range = "P2:R" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_5_tv_header)
                tab.tableStyleInfo = style
                worksheet4.add_table(tab)

                (max_row, max_col) = table_6_tv.shape
                max_row = retd(max_row)
                cell_range = "T2:V" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_6_tv_header)
                tab.tableStyleInfo = style
                worksheet4.add_table(tab)
                
                (max_row, max_col) = table_7_tv.shape
                max_row = retd(max_row)
                cell_range = "X2:Z" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_7_tv_header)
                tab.tableStyleInfo = style
                worksheet4.add_table(tab)

                (max_row, max_col) = table_8_tv.shape
                max_row = retd(max_row)
                cell_range = "AB2:AD" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_8_tv_header)
                tab.tableStyleInfo = style
                worksheet4.add_table(tab)

                (max_row, max_col) = table_9_tv.shape
                max_row = retd(max_row)
                cell_range = "AF2:AH" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_9_tv_header)
                tab.tableStyleInfo = style
                worksheet4.add_table(tab)

                (max_row, max_col) = table_10_tv.shape
                max_row = retd(max_row)
                cell_range = "AJ2:AL" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_10_tv_header)
                tab.tableStyleInfo = style
                worksheet4.add_table(tab)

#**********************************************************************************************************************************************************
#   TABLE's FORMITTING AND AUTO CELL ALIGNMENT FOR WORKSHEET5
#**********************************************************************************************************************************************************


                for idx, col in enumerate(worksheet5, 1):
                        worksheet5.column_dimensions[get_column_letter(idx)].auto_size = True


                (max_row, max_col) = table_1_ts.shape
                max_row = retd(max_row)
                cell_range = "A2:B" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_1_ts_header)
                tab.tableStyleInfo = style
                worksheet5.add_table(tab)

                
                (max_row, max_col) = table_2_ts.shape
                max_row = retd(max_row)
                cell_range = "D2:F" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_2_ts_header)
                tab.tableStyleInfo = style
                worksheet5.add_table(tab)

                (max_row, max_col) = table_3_ts.shape
                max_row = retd(max_row)
                cell_range = "H2:J" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_3_ts_header)
                tab.tableStyleInfo = style
                worksheet5.add_table(tab)        

                (max_row, max_col) = table_4_ts.shape
                max_row = retd(max_row)
                cell_range = "L2:N" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_4_ts_header)
                tab.tableStyleInfo = style
                worksheet5.add_table(tab)        

                (max_row, max_col) = table_5_ts.shape
                max_row = retd(max_row)
                cell_range = "P2:R" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_5_ts_header)
                tab.tableStyleInfo = style
                worksheet5.add_table(tab)

                (max_row, max_col) = table_6_ts.shape
                max_row = retd(max_row)
                cell_range = "T2:V" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_6_ts_header)
                tab.tableStyleInfo = style
                worksheet5.add_table(tab)
                
                (max_row, max_col) = table_7_ts.shape
                max_row = retd(max_row)
                cell_range = "X2:Z" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_7_ts_header)
                tab.tableStyleInfo = style
                worksheet5.add_table(tab)

                (max_row, max_col) = table_8_ts.shape
                max_row = retd(max_row)
                cell_range = "AB2:AD" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_8_ts_header)
                tab.tableStyleInfo = style
                worksheet5.add_table(tab)

                (max_row, max_col) = table_9_ts.shape
                max_row = retd(max_row)
                cell_range = "AF2:AH" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_9_ts_header)
                tab.tableStyleInfo = style
                worksheet5.add_table(tab)

                (max_row, max_col) = table_10_ts.shape
                max_row = retd(max_row)
                cell_range = "AJ2:AL" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_10_ts_header)
                tab.tableStyleInfo = style
                worksheet5.add_table(tab)

#**********************************************************************************************************************************************************
#   TABLE's FORMITTING AND AUTO CELL ALIGNMENT FOR WORKSHEET6
#**********************************************************************************************************************************************************


                for idx, col in enumerate(worksheet6, 1):
                        worksheet6.column_dimensions[get_column_letter(idx)].auto_size = True


                (max_row, max_col) = table_1_tp.shape
                max_row = retd(max_row)
                cell_range = "A2:B" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_1_tp_header)
                tab.tableStyleInfo = style
                worksheet6.add_table(tab)

                
                (max_row, max_col) = table_2_tp.shape
                max_row = retd(max_row)
                cell_range = "D2:F" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_2_tp_header)
                tab.tableStyleInfo = style
                worksheet6.add_table(tab)

                (max_row, max_col) = table_3_tp.shape
                max_row = retd(max_row)
                cell_range = "H2:J" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_3_tp_header)
                tab.tableStyleInfo = style
                worksheet6.add_table(tab)        

                (max_row, max_col) = table_4_tp.shape
                max_row = retd(max_row)
                cell_range = "L2:N" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_4_tp_header)
                tab.tableStyleInfo = style
                worksheet6.add_table(tab)        

                (max_row, max_col) = table_5_tp.shape
                max_row = retd(max_row)
                cell_range = "P2:R" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_5_tp_header)
                tab.tableStyleInfo = style
                worksheet6.add_table(tab)

                (max_row, max_col) = table_6_tp.shape
                max_row = retd(max_row)
                cell_range = "T2:V" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_6_tp_header)
                tab.tableStyleInfo = style
                worksheet6.add_table(tab)
                
                (max_row, max_col) = table_7_tp.shape
                max_row = retd(max_row)
                cell_range = "X2:Z" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_7_tp_header)
                tab.tableStyleInfo = style
                worksheet6.add_table(tab)

                (max_row, max_col) = table_8_tp.shape
                max_row = retd(max_row)
                cell_range = "AB2:AD" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_8_tp_header)
                tab.tableStyleInfo = style
                worksheet6.add_table(tab)

                (max_row, max_col) = table_9_tp.shape
                max_row = retd(max_row)
                cell_range = "AF2:AH" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_9_tp_header)
                tab.tableStyleInfo = style
                worksheet6.add_table(tab)

                (max_row, max_col) = table_10_tp.shape
                max_row = retd(max_row)
                cell_range = "AJ2:AL" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_10_tp_header)
                tab.tableStyleInfo = style
                worksheet6.add_table(tab)

#**********************************************************************************************************************************************************
#   TABLE's FORMITTING AND AUTO CELL ALIGNMENT FOR WORKSHEET7
#**********************************************************************************************************************************************************


                for idx, col in enumerate(worksheet7, 1):
                        worksheet7.column_dimensions[get_column_letter(idx)].auto_size = True


                (max_row, max_col) = table_1_cr.shape
                max_row = retd(max_row)
                cell_range = "A2:B" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_1_cr_header)
                tab.tableStyleInfo = style
                worksheet7.add_table(tab)

                
                (max_row, max_col) = table_2_cr.shape
                max_row = retd(max_row)
                cell_range = "D2:E" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_2_cr_header)
                tab.tableStyleInfo = style
                worksheet7.add_table(tab)

                (max_row, max_col) = table_3_cr.shape
                max_row = retd(max_row)
                cell_range = "G2:H" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_3_cr_header)
                tab.tableStyleInfo = style
                worksheet7.add_table(tab)        

                (max_row, max_col) = table_4_cr.shape
                max_row = retd(max_row)
                cell_range = "J2:K" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_4_cr_header)
                tab.tableStyleInfo = style
                worksheet7.add_table(tab)        

                (max_row, max_col) = table_5_cr.shape
                max_row = retd(max_row)
                cell_range = "M2:N" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_5_cr_header)
                tab.tableStyleInfo = style
                worksheet7.add_table(tab)

                (max_row, max_col) = table_6_cr.shape
                max_row = retd(max_row)
                cell_range = "P2:Q" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_6_cr_header)
                tab.tableStyleInfo = style
                worksheet7.add_table(tab)
                
                (max_row, max_col) = table_7_cr.shape
                max_row = retd(max_row)
                cell_range = "S2:T" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_7_cr_header)
                tab.tableStyleInfo = style
                worksheet7.add_table(tab)

                (max_row, max_col) = table_8_cr.shape
                max_row = retd(max_row)
                cell_range = "V2:W" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_8_cr_header)
                tab.tableStyleInfo = style
                worksheet7.add_table(tab)

                (max_row, max_col) = table_9_cr.shape
                max_row = retd(max_row)
                cell_range = "Y2:Z" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_9_cr_header)
                tab.tableStyleInfo = style
                worksheet7.add_table(tab)

                (max_row, max_col) = table_10_cr.shape
                max_row = retd(max_row)
                cell_range = "AB2:AC" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_10_cr_header)
                tab.tableStyleInfo = style
                worksheet7.add_table(tab)

                (max_row, max_col) = table_11_cr.shape
                max_row = retd(max_row)
                cell_range = "AE2:AF" + str(max_row + 2)
                tab = Table(ref=cell_range, displayName= table_11_cr_header)
                tab.tableStyleInfo = style
                worksheet7.add_table(tab)

#**********************************************************************************************************************************************************
#   THIS LOOP IS FOR ALL WORKSHEET's WIDTH ADJUSTMENT
#**********************************************************************************************************************************************************

# This loop is for worksheet table's width adjustment

                for i in range(len(wb.worksheets)):
                    wst = wb.worksheets[i]
                    tables = wst.tables.items()
                    for table, rng in tables:
                        wst[rng.split(':')[0][:-1] + str(int(rng.split(':')[0][-1]) - 1)] = table
                        
            


#**********************************************************************************************************************************************************
#   THIS FUCTION WILL CREATES ALL TABLE_*_AA TABLES FOR "Approved Actuals by Skill" WORKBOOK
#**********************************************************************************************************************************************************

def gen_table_aa():

# Table__1_Approved_Hours_For_Actuals query

        global table_1_aa
        
        table_1_aa = mysql('''SELECT Raw_Data.WeekendDate as WeekendDate, sum(Raw_Data.Hours) as Expr1001
                        FROM Raw_Data LEFT JOIN Resource_Master
                        ON upper(trim(Raw_Data.ResourceName)) = upper(trim(Resource_Master.ResourceName))
                        WHERE Raw_Data.ResourceName not like '2%'
                        AND Raw_Data.Status = 'Approved'
                        GROUP BY Raw_Data.WeekendDate;''')




# Intermediate temp table to extract all table_*_aa table results

        global Temp_Table

        Temp_Table = mysql('''SELECT Raw_Data.WeekendDate as WeekendDate, Raw_Data.Hours,  Resource_Master.Skill as SkillExp, Raw_Data.ResourceName AS ResourceName
                                FROM Raw_Data LEFT JOIN Resource_Master
                                ON upper(trim(Raw_Data.ResourceName)) = upper(trim(Resource_Master.ResourceName))
                                WHERE Raw_Data.ResourceName not like '2%'
                                AND Raw_Data.Status = 'Approved';''')

        Temp_Table['SkillExp'] = Temp_Table.apply(lambda row: skillname(str(row['ResourceName'])).upper().strip()
                                 if pd.isnull(row['SkillExp']) else row['SkillExp'].upper().strip(), axis=1)

# Table_2_Approved_Mainframe_Hours_For_Actuals query

        global table_2_aa

        table_2_aa = mysql('''SELECT WeekendDate, sum(Hours) as Expr1001, SkillExp
                        FROM Temp_Table
                        WHERE SkillExp = 'MAINFRAME'
                        GROUP BY WeekendDate, SkillExp;''')

       
# Table__3_Approved_DotNet_Hours_For_Actuals query   
        
        global table_3_aa

        table_3_aa = mysql('''SELECT WeekendDate, sum(Hours) as Expr1001, SkillExp
                        FROM Temp_Table
                        WHERE SkillExp = 'DOT NET'
                        GROUP BY WeekendDate, SkillExp;''')

# Table__4_Approved_Java_Hours_For_Actuals query

        global table_4_aa

        table_4_aa = mysql('''SELECT WeekendDate, sum(Hours) as Expr1001, SkillExp
                        FROM Temp_Table
                        WHERE SkillExp = 'JAVA'
                        GROUP BY WeekendDate, SkillExp;''')

# Table__5_Approved_Dba_Hours_For_Actuals query

        global table_5_aa

        table_5_aa = mysql('''SELECT WeekendDate, sum(Hours) as Expr1001, SkillExp
                        FROM Temp_Table
                        WHERE SkillExp = 'DBA'
                        GROUP BY WeekendDate, SkillExp;''')

# Table__6_Approved_Pm_Hours_For_Actuals query

        global table_6_aa

        table_6_aa = mysql('''SELECT WeekendDate, sum(Hours) as Expr1001, SkillExp
                        FROM Temp_Table
                        WHERE SkillExp = 'PM'
                        GROUP BY WeekendDate, SkillExp;''')

# Table__7_Approved_Pgm_Hours_For_Actuals query

        global table_7_aa

        table_7_aa = mysql('''SELECT WeekendDate, sum(Hours) as Expr1001, SkillExp
                        FROM Temp_Table
                        WHERE SkillExp = 'SME'
                        GROUP BY WeekendDate, SkillExp;''')

# Table__8_Approved_Other_Hours_For_Actuals query


        global table_8_aa

        table_8_aa = mysql('''SELECT WeekendDate, sum(Hours) as Expr1001, 'OTHER' as SkillExp
                        FROM Temp_Table
                        WHERE SkillExp NOT IN ('MAINFRAME', 'DOT NET', 'JAVA', 'DBA', 'ORACLE DBA', 'PM', 'SME', 'BA', 'DOC SPL')
                        GROUP BY WeekendDate, 'OTHER';''')


# Table__9_Approved_Ba_Hours_For_Actuals query


        global table_9_aa

        table_9_aa = mysql('''SELECT WeekendDate, sum(Hours) as Expr1001, SkillExp
                        FROM Temp_Table
                        WHERE SkillExp = 'BA'
                        GROUP BY WeekendDate, SkillExp;''')

# Table__10_Approved_Doc_Hours_For_Actuals query

        global table_10_aa

        table_10_aa = mysql('''SELECT WeekendDate, sum(Hours) as Expr1001, SkillExp
                        FROM Temp_Table
                        WHERE SkillExp = 'DOC SPL'
                        GROUP BY WeekendDate, SkillExp;''')


#**********************************************************************************************************************************************************
#   THIS FUCTION WILL CREATES ALL TABLE_*_TV TABLES FOR "Total Vacations By Skill" WORKBOOK
#**********************************************************************************************************************************************************

def gen_table_tv():


# Table__1_Vacation_Hours_For_Actuals query

        global table_1_tv
        
        table_1_tv = mysql('''SELECT Raw_Data.WeekendDate as WeekendDate, sum(Raw_Data.Hours) as Expr1001
                        FROM Raw_Data LEFT JOIN Resource_Master
                        ON upper(trim(Raw_Data.ResourceName)) = upper(trim(Resource_Master.ResourceName))
                        WHERE Raw_Data.ResourceName not like '2%'
                        AND Raw_Data.Status = 'Vacation'
                        GROUP BY Raw_Data.WeekendDate;''')

# Intermediate temp table to extract all table_*_tv table results

        
        Temp_Table = mysql('''SELECT Raw_Data.WeekendDate, Raw_Data.Hours,  Resource_Master.Skill as SkillExp, Raw_Data.ResourceName AS ResourceName
                                FROM Raw_Data LEFT JOIN Resource_Master
                                ON upper(trim(Raw_Data.ResourceName)) = upper(trim(Resource_Master.ResourceName))
                                WHERE Raw_Data.ResourceName not like '2%'
                                AND Raw_Data.Status = 'Vacation';''')

        Temp_Table['SkillExp'] = Temp_Table.apply(lambda row: skillname(str(row['ResourceName'])).upper().strip()
                                 if pd.isnull(row['SkillExp']) else row['SkillExp'].upper().strip(), axis=1)


# Table_2_Vacation_Mainframe_Hours_For_Actuals query

        global table_2_tv

        table_2_tv = mysql('''SELECT WeekendDate, sum(Hours) as Expr1001, SkillExp
                        FROM Temp_Table
                        WHERE SkillExp = 'MAINFRAME'
                        GROUP BY WeekendDate, SkillExp;''')

       
# Table__3_Vacation_DotNet_Hours_For_Actuals query   
        
        global table_3_tv

        table_3_tv = mysql('''SELECT WeekendDate, sum(Hours) as Expr1001, SkillExp
                        FROM Temp_Table
                        WHERE SkillExp = 'DOT NET'
                        GROUP BY WeekendDate, SkillExp;''')

# Table__4_Vacation_Java_Hours_For_Actuals query

        global table_4_tv

        table_4_tv = mysql('''SELECT WeekendDate, sum(Hours) as Expr1001, SkillExp
                        FROM Temp_Table
                        WHERE SkillExp = 'JAVA'
                        GROUP BY WeekendDate, SkillExp;''')

# Table__5_Vacation_Dba_Hours_For_Actuals query

        global table_5_tv

        table_5_tv = mysql('''SELECT WeekendDate, sum(Hours) as Expr1001, SkillExp
                        FROM Temp_Table
                        WHERE SkillExp = 'DBA'
                        GROUP BY WeekendDate, SkillExp;''')

# Table__6_Vacation_Pm_Hours_For_Actuals query

        global table_6_tv

        table_6_tv = mysql('''SELECT WeekendDate, sum(Hours) as Expr1001, SkillExp
                        FROM Temp_Table
                        WHERE SkillExp = 'PM'
                        GROUP BY WeekendDate, SkillExp;''')

# Table__7_Vacation_Pgm_Hours_For_Actuals query

        global table_7_tv

        table_7_tv = mysql('''SELECT WeekendDate, sum(Hours) as Expr1001, SkillExp
                        FROM Temp_Table
                        WHERE SkillExp = 'SME'
                        GROUP BY WeekendDate, SkillExp;''')

# Table__8_Vacation_Other_Hours_For_Actuals query


        global table_8_tv

        table_8_tv = mysql('''SELECT WeekendDate, sum(Hours) as Expr1001, 'OTHER' as SkillExp
                        FROM Temp_Table
                        WHERE SkillExp NOT IN ('MAINFRAME', 'DOT NET', 'JAVA', 'DBA', 'ORACLE DBA', 'PM', 'SME', 'BA', 'DOC SPL')
                        GROUP BY WeekendDate, 'OTHER';''')


# Table__9_Vacation_Ba_Hours_For_Actuals query


        global table_9_tv

        table_9_tv = mysql('''SELECT WeekendDate, sum(Hours) as Expr1001, SkillExp
                        FROM Temp_Table
                        WHERE SkillExp = 'BA'
                        GROUP BY WeekendDate, SkillExp;''')

# Table__10_Vacation_Doc_Hours_For_Actuals query

        global table_10_tv

        table_10_tv = mysql('''SELECT WeekendDate, sum(Hours) as Expr1001, SkillExp
                        FROM Temp_Table
                        WHERE SkillExp = 'DOC SPL'
                        GROUP BY WeekendDate, SkillExp;''')


#**********************************************************************************************************************************************************
#   THIS FUCTION WILL CREATES ALL TABLE_*_TS TABLES FOR "Total Submitted By Skill" WORKBOOK
#**********************************************************************************************************************************************************

def gen_table_ts():
    

# Table__1_Submitted_Hours_For_Actuals query

        global table_1_ts
        
        table_1_ts = mysql('''SELECT Raw_Data.WeekendDate as WeekendDate, sum(Raw_Data.Hours) as Expr1001
                        FROM Raw_Data LEFT JOIN Resource_Master
                        ON upper(trim(Raw_Data.ResourceName)) = upper(trim(Resource_Master.ResourceName))
                        WHERE Raw_Data.ResourceName not like '2%'
                        AND (Raw_Data.Status = 'ROM' OR
                        Raw_Data.Status = 'Submitted')
                        GROUP BY Raw_Data.WeekendDate;''')

# Intermediate temp table to extract all table_*_ts table results

        
        Temp_Table = mysql('''SELECT Raw_Data.WeekendDate, Raw_Data.Hours,  Resource_Master.Skill as SkillExp, Raw_Data.ResourceName AS ResourceName
                                FROM Raw_Data LEFT JOIN Resource_Master
                                ON upper(trim(Raw_Data.ResourceName)) = upper(trim(Resource_Master.ResourceName))
                                WHERE Raw_Data.ResourceName not like '2%'
                                AND (Raw_Data.Status = 'ROM' OR
                                     Raw_Data.Status = 'Submitted');''')

        Temp_Table['SkillExp'] = Temp_Table.apply(lambda row: skillname(str(row['ResourceName'])).upper().strip()
                                 if pd.isnull(row['SkillExp']) else row['SkillExp'].upper().strip(), axis=1)


# Table_2_Submitted_Mainframe_Hours_For_Actuals query

        global table_2_ts

        table_2_ts = mysql('''SELECT WeekendDate, sum(Hours) as Expr1001, SkillExp
                        FROM Temp_Table
                        WHERE SkillExp = 'MAINFRAME'
                        GROUP BY WeekendDate, SkillExp;''')

       
# Table__3_Submitted_DotNet_Hours_For_Actuals query   
        
        global table_3_ts

        table_3_ts = mysql('''SELECT WeekendDate, sum(Hours) as Expr1001, SkillExp
                        FROM Temp_Table
                        WHERE SkillExp = 'DOT NET'
                        GROUP BY WeekendDate, SkillExp;''')

# Table__4_Submitted_Java_Hours_For_Actuals query

        global table_4_ts

        table_4_ts = mysql('''SELECT WeekendDate, sum(Hours) as Expr1001, SkillExp
                        FROM Temp_Table
                        WHERE SkillExp = 'JAVA'
                        GROUP BY WeekendDate, SkillExp;''')

# Table__5_Submitted_Dba_Hours_For_Actuals query

        global table_5_ts

        table_5_ts = mysql('''SELECT WeekendDate, sum(Hours) as Expr1001, SkillExp
                        FROM Temp_Table
                        WHERE SkillExp = 'DBA'
                        GROUP BY WeekendDate, SkillExp;''')

# Table__6_Submitted_Pm_Hours_For_Actuals query

        global table_6_ts

        table_6_ts = mysql('''SELECT WeekendDate, sum(Hours) as Expr1001, SkillExp
                        FROM Temp_Table
                        WHERE SkillExp = 'PM'
                        GROUP BY WeekendDate, SkillExp;''')

# Table__7_Submitted_Pgm_Hours_For_Actuals query

        global table_7_ts

        table_7_ts = mysql('''SELECT WeekendDate, sum(Hours) as Expr1001, SkillExp
                        FROM Temp_Table
                        WHERE SkillExp = 'SME'
                        GROUP BY WeekendDate, SkillExp;''')

# Table__8_Submitted_Other_Hours_For_Actuals query


        global table_8_ts

        table_8_ts = mysql('''SELECT WeekendDate, sum(Hours) as Expr1001, 'OTHER' as SkillExp
                        FROM Temp_Table
                        WHERE SkillExp NOT IN ('MAINFRAME', 'DOT NET', 'JAVA', 'DBA', 'ORACLE DBA', 'PM', 'SME', 'BA', 'DOC SPL')
                        GROUP BY WeekendDate, 'OTHER';''')


# Table__9_Submitted_Ba_Hours_For_Actuals query


        global table_9_ts

        table_9_ts = mysql('''SELECT WeekendDate, sum(Hours) as Expr1001, SkillExp
                        FROM Temp_Table
                        WHERE SkillExp = 'BA'
                        GROUP BY WeekendDate, SkillExp;''')

# Table__10_Submitted_Doc_Hours_For_Actuals query

        global table_10_ts

        table_10_ts = mysql('''SELECT WeekendDate, sum(Hours) as Expr1001, SkillExp
                        FROM Temp_Table
                        WHERE SkillExp = 'DOC SPL'
                        GROUP BY WeekendDate, SkillExp;''')


#**********************************************************************************************************************************************************
#   THIS FUCTION WILL CREATES ALL TABLE_*_TP TABLES FOR "Total Placeholder By Skill" WORKBOOK
#**********************************************************************************************************************************************************

def gen_table_tp():


# Table__1_Total_Hours_For_Placeholders query

        global table_1_tp
        
        table_1_tp = mysql('''SELECT Raw_Data.WeekendDate as WeekendDate, sum(Raw_Data.Hours) as Expr1001
                        FROM Raw_Data LEFT JOIN Resource_Master
                        ON upper(trim(Raw_Data.ResourceName)) = upper(trim(Resource_Master.ResourceName))
                        WHERE Raw_Data.ResourceName like '2%'
                        GROUP BY Raw_Data.WeekendDate;''')

# Intermediate temp table to extract all table_*_tp table results

        
        Temp_Table = mysql('''SELECT Raw_Data.WeekendDate, Raw_Data.Hours,  Resource_Master.Skill as SkillExp, Raw_Data.ResourceName AS ResourceName
                                FROM Raw_Data LEFT JOIN Resource_Master
                                ON upper(trim(Raw_Data.ResourceName)) = upper(trim(Resource_Master.ResourceName))
                                WHERE Raw_Data.ResourceName like '2%';''')
                                

        Temp_Table['SkillExp'] = Temp_Table.apply(lambda row: skillname(str(row['ResourceName'])).upper().strip()
                                 if pd.isnull(row['SkillExp']) else row['SkillExp'].upper().strip(), axis=1)


# Table__2_Total_Mainframe_Hours_For_Placeholders query

        global table_2_tp

        table_2_tp = mysql('''SELECT WeekendDate, sum(Hours) as Expr1001, SkillExp
                        FROM Temp_Table
                        WHERE SkillExp = 'MAINFRAME'
                        GROUP BY WeekendDate, SkillExp;''')

       
# Table__3_Total_DotNet_Hours_For_Placeholders query   
        
        global table_3_tp

        table_3_tp = mysql('''SELECT WeekendDate, sum(Hours) as Expr1001, SkillExp
                        FROM Temp_Table
                        WHERE SkillExp = 'DOT NET'
                        GROUP BY WeekendDate, SkillExp;''')

# Table__4_Total_Java_Hours_For_Placeholders query

        global table_4_tp

        table_4_tp = mysql('''SELECT WeekendDate, sum(Hours) as Expr1001, SkillExp
                        FROM Temp_Table
                        WHERE SkillExp = 'JAVA'
                        GROUP BY WeekendDate, SkillExp;''')

# Table__5_Total_Dba_Hours_For_Placeholders query

        global table_5_tp

        table_5_tp = mysql('''SELECT WeekendDate, sum(Hours) as Expr1001, SkillExp
                        FROM Temp_Table
                        WHERE SkillExp = 'DBA'
                        GROUP BY WeekendDate, SkillExp;''')

# Table__6_Total_Pm_Hours_For_Placeholders query

        global table_6_tp

        table_6_tp = mysql('''SELECT WeekendDate, sum(Hours) as Expr1001, SkillExp
                        FROM Temp_Table
                        WHERE SkillExp = 'PM'
                        GROUP BY WeekendDate, SkillExp;''')

# Table__7_Total_Pgm_Hours_For_Placeholders query

        global table_7_tp

        table_7_tp = mysql('''SELECT WeekendDate, sum(Hours) as Expr1001, SkillExp
                        FROM Temp_Table
                        WHERE SkillExp = 'SME'
                        GROUP BY WeekendDate, SkillExp;''')

# Table__8_Total_Other_Hours_For_Placeholders query


        global table_8_tp

        table_8_tp = mysql('''SELECT WeekendDate, sum(Hours) as Expr1001, 'OTHER' as SkillExp
                        FROM Temp_Table
                        WHERE SkillExp NOT IN ('MAINFRAME', 'DOT NET', 'JAVA', 'DBA', 'ORACLE DBA', 'PM', 'SME', 'BA', 'DOC SPL')
                        GROUP BY WeekendDate, 'OTHER';''')


# Table__9_Total_Ba_Hours_For_Placeholders query


        global table_9_tp

        table_9_tp = mysql('''SELECT WeekendDate, sum(Hours) as Expr1001, SkillExp
                        FROM Temp_Table
                        WHERE SkillExp = 'BA'
                        GROUP BY WeekendDate, SkillExp;''')

# Table__10_Total_Doc_Hours_For_Placeholders query

        global table_10_tp

        table_10_tp = mysql('''SELECT WeekendDate, sum(Hours) as Expr1001, SkillExp
                        FROM Temp_Table
                        WHERE SkillExp = 'DOC SPL'
                        GROUP BY WeekendDate, SkillExp;''')

#**********************************************************************************************************************************************************
#   THIS FUCTION WILL CREATES ALL TABLE_*_CR TABLES FOR "Covered Resource Counts" WORKBOOK
#**********************************************************************************************************************************************************

def gen_table_cr():

# Table__1_Covered_Resource_Count query

        global table_1_cr
        
        table_1_cr = mysql('''SELECT T1.WeekendDate, count(*) as Covered_Resource_Count
                           FROM
                            (SELECT DISTINCT Raw_Data.WeekendDate, upper(Raw_Data.ResourceName)
                                FROM Raw_Data LEFT JOIN Resource_Master
                                ON upper(trim(Raw_Data.ResourceName)) = upper(trim(Resource_Master.ResourceName))
                                WHERE Raw_Data.ResourceName not like '2%') AS T1
                                GROUP BY T1.WeekendDate;''')

# Intermediate temp table to extract all table_*_cr table results

        
        Temp_Table = mysql('''SELECT DISTINCT Raw_Data.WeekendDate, upper(Raw_Data.ResourceName) AS ResourceName, Resource_Master.Skill AS SkillExp
                                FROM Raw_Data LEFT JOIN Resource_Master
                                ON upper(trim(Raw_Data.ResourceName)) = upper(trim(Resource_Master.ResourceName))
                                WHERE Raw_Data.ResourceName not like '2%';''')

        
        
        Temp_Table['SkillExp'] = Temp_Table.apply(lambda row: skillname(str(row['ResourceName'])).upper().strip()
                                 if pd.isnull(row['SkillExp']) else row['SkillExp'].upper().strip(), axis=1)
        
        

# Table__2_Covered_Mainframe_Resource_Count query

        global table_2_cr

        table_2_cr = mysql('''SELECT WeekendDate, count(*) as Covered_Resource_Count
                        FROM Temp_Table
                        WHERE SkillExp = 'MAINFRAME'
                        GROUP BY WeekendDate;''')

       
# Table__3_Covered_DotNet_Resource_Count query   
        
        global table_3_cr

        table_3_cr = mysql('''SELECT WeekendDate, count(*) as Covered_Resource_Count
                        FROM Temp_Table
                        WHERE SkillExp = 'DOT NET'
                        GROUP BY WeekendDate;''')
        

# Table__4_Covered_Java_Resource_Count query

        global table_4_cr

        table_4_cr = mysql('''SELECT WeekendDate, count(*) as Covered_Resource_Count
                        FROM Temp_Table
                        WHERE SkillExp = 'JAVA'
                        GROUP BY WeekendDate;''')
        

# Table__5_Covered_Dba_Resource_Count query

        global table_5_cr

        table_5_cr = mysql('''SELECT WeekendDate, count(*) as Covered_Resource_Count
                        FROM Temp_Table
                        WHERE SkillExp = 'DBA'
                        GROUP BY WeekendDate;''')
                           

# Table__6_Covered_Pm_Resource_Count query

        global table_6_cr

        table_6_cr = mysql('''SELECT WeekendDate, count(*) as Covered_Resource_Count
                        FROM Temp_Table
                        WHERE SkillExp = 'PM'
                        GROUP BY WeekendDate;''')

# Table__7_Covered_Sme_Resource_Count query

        global table_7_cr

        table_7_cr = mysql('''SELECT WeekendDate, count(*) as Covered_Resource_Count
                        FROM Temp_Table
                        WHERE SkillExp = 'SME'
                        GROUP BY WeekendDate;''')                           

# Table__8_Covered_Other_Resource_Count query


        global table_8_cr

        table_8_cr = mysql('''SELECT WeekendDate, count(*) as Covered_Resource_Count
                        FROM Temp_Table
                        WHERE SkillExp NOT IN ('MAINFRAME', 'DOT NET', 'JAVA', 'DBA', 'ORACLE DBA', 'PM', 'SME', 'BA', 'DOC SPL')
                        GROUP BY WeekendDate;''')


# Table__09_Covered_Ba_Resource_Count query

        global table_9_cr

        table_9_cr = mysql('''SELECT WeekendDate, count(*) as Covered_Resource_Count
                        FROM Temp_Table
                        WHERE SkillExp = 'BA'
                        GROUP BY WeekendDate;''')
                           

# Table__10_Covered_Doc_Resource_Count query

        global table_10_cr

        table_10_cr = mysql('''SELECT WeekendDate, count(*) as Covered_Resource_Count
                        FROM Temp_Table
                        WHERE SkillExp = 'DOC SPL'
                        GROUP BY WeekendDate;''')
                           


# Table__11_Resource_Master_Counts_By_Type query


        Temp_Table = mysql('''SELECT Resource_Master.Skill as SkillExp FROM Resource_Master;''')
        
        Temp_Table['SkillExp'] = Temp_Table['SkillExp'].apply(lambda row: SkillCategory(str(row)))

        
        global table_11_cr

        table_11_cr = mysql('''SELECT T1.*
                                FROM ( SELECT SkillExp as SkillCat, count(*) as Resource_Count from Temp_Table
                                    GROUP BY SkillExp UNION
                                            SELECT "TOTAL", COUNT(*) as Resource_Count FROM Resource_Master) AS T1''')

        

#**********************************************************************************************************************************************************
#   START SCRIPT ONCE HIT ON START BUTTON
#**********************************************************************************************************************************************************

def start_app():

#FOR REFERENCE IF USE showinfo it will return => ok , showwarning => ok, showerror => ok, askquestion => 1 or 0, askokcancel => 1 or 0, askyesno => 1 or 0

        if read_raw_data == 'y' and read_resource_master == 'y' :

            response = messagebox.askyesno("Information" , "Select Yes to Proceed/ No to Exit")

            if response == 1:
                gen_table_aa()
                gen_table_tv()
                gen_table_ts()
                gen_table_tp()
                gen_table_cr()
                create_output()
                close_dialog_box()
    
                messagebox.showinfo("Information" , "Output Generated into same folder!")

            else:
                close_dialog_box()
                messagebox.showinfo("Information" , "Thank You! Please Try Later!")
        else:

            messagebox.showinfo("Information" , "Please Select Both File to Start!")


root.mainloop()



